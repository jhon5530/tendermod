"""
Parse Gold Standard Excel files produced by Claude Cowork.

Expected format (based on Requerimientos_Pliego_Ciberseguridad.xlsx):
  Sheet "Requerimientos": ID | Categoría | Tipo | Sección Pliego | Requerimiento | Descripción | ...
  Sheet "Indicadores":    tables with rows like (Indicador, Fórmula, Umbral exigido, ...)
  Sheet "Experiencia":    rows describing experience segments / requirements
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

CATEGORY_MAP: dict[str, str] = {
    "jurídico": "JURIDICO",
    "juridico": "JURIDICO",
    "técnico": "TECNICO",
    "tecnico": "TECNICO",
    "documentación": "DOCUMENTACION",
    "documentacion": "DOCUMENTACION",
    "financiero": "FINANCIERO_OTRO",
    "experiencia": "EXPERIENCIA",
    "garantía": "GARANTIA",
    "garantia": "GARANTIA",
    "evaluación": "EVALUACION",
    "evaluacion": "EVALUACION",
    "capacidad": "CAPACIDAD",
    "otro": "OTRO",
    "causal": "CAUSAL_RECHAZO",
}

TYPE_MAP: dict[str, str] = {
    "habilitante": "HABILITANTE",
    "puntuable": "PUNTUABLE",
    "documental": "DOCUMENTAL",
    "garantía": "GARANTIA",
    "garantia": "GARANTIA",
}


@dataclass
class GoldRequirement:
    id: int
    categoria: str          # normalized to JURIDICO, TECNICO, etc.
    tipo: str               # normalized to HABILITANTE, PUNTUABLE, etc.
    seccion: str
    nombre: str
    descripcion: str        # text used for semantic comparison
    documento_formato: str = ""
    categoria_raw: str = ""
    tipo_raw: str = ""


@dataclass
class GoldIndicator:
    nombre: str
    formula: str = ""
    umbral_raw: str = ""    # raw string like "≥ 1,13"
    umbral_valor: Optional[float] = None
    umbral_condicion: str = ""   # "gte", "lte", "gt", "lt"


@dataclass
class GoldExperienceSegment:
    nombre: str
    smmlv_minimos: Optional[float] = None
    descripcion: str = ""


@dataclass
class GoldStandard:
    pdf_name: str
    requirements: list[GoldRequirement] = field(default_factory=list)
    indicators: list[GoldIndicator] = field(default_factory=list)
    experience: list[GoldExperienceSegment] = field(default_factory=list)


def _normalize_category(raw: str) -> str:
    if not raw:
        return "OTRO"
    key = raw.strip().lower()
    for k, v in CATEGORY_MAP.items():
        if k in key:
            return v
    return "OTRO"


def _normalize_type(raw: str) -> str:
    if not raw:
        return "NO_ESPECIFICADO"
    key = raw.strip().lower()
    for k, v in TYPE_MAP.items():
        if k in key:
            return v
    return "NO_ESPECIFICADO"


def _parse_threshold(raw: str) -> tuple[Optional[float], str]:
    """Parse threshold string like '≥ 1,13' or '≤ 0,84' into (value, condition)."""
    if not raw:
        return None, ""
    raw = str(raw).strip()
    cond = ""
    if "≥" in raw or ">=" in raw or "mayor" in raw.lower() or "mínimo" in raw.lower():
        cond = "gte"
    elif "≤" in raw or "<=" in raw or "menor" in raw.lower() or "máximo" in raw.lower():
        cond = "lte"
    elif ">" in raw:
        cond = "gt"
    elif "<" in raw:
        cond = "lt"
    # extract numeric part
    num_str = re.sub(r"[^\d,\.]", "", raw).replace(",", ".")
    try:
        return float(num_str), cond
    except (ValueError, TypeError):
        return None, cond


def _find_header_row(ws, candidate_cols: list[str]) -> Optional[int]:
    """Return 1-based row index where at least 2 of candidate_cols appear as headers."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip().lower() if c else "" for c in row]
        hits = sum(1 for col in candidate_cols if any(col in cell for cell in cells))
        if hits >= 2:
            return row_idx
    return None


def _col_index(header_row: list, candidates: list[str]) -> int:
    """Return 0-based index of first column whose lowercase header matches any candidate."""
    lower = [str(c).strip().lower() if c else "" for c in header_row]
    for i, cell in enumerate(lower):
        for cand in candidates:
            if cand in cell:
                return i
    return -1


def _parse_requirements_sheet(ws) -> list[GoldRequirement]:
    reqs: list[GoldRequirement] = []
    header_row_idx = _find_header_row(ws, ["categoría", "categoria", "tipo", "descripción", "descripcion"])
    if header_row_idx is None:
        return reqs

    rows = list(ws.iter_rows(values_only=True))
    header = rows[header_row_idx - 1]

    col_id = _col_index(header, ["id", "#"])
    col_cat = _col_index(header, ["categoría", "categoria"])
    col_tipo = _col_index(header, ["tipo"])
    col_sec = _col_index(header, ["sección", "seccion", "numeral"])
    col_nombre = _col_index(header, ["requerimiento", "nombre", "requisito"])
    col_desc = _col_index(header, ["descripción", "descripcion"])
    col_doc = _col_index(header, ["documento", "formato"])

    def cell(row, idx):
        if idx < 0 or idx >= len(row):
            return ""
        return str(row[idx]).strip() if row[idx] is not None else ""

    req_id = 1
    for row in rows[header_row_idx:]:
        if all(c is None for c in row):
            continue
        cat_raw = cell(row, col_cat)
        if not cat_raw:
            continue
        nombre = cell(row, col_nombre)
        desc = cell(row, col_desc)
        if not nombre and not desc:
            continue
        # use Descripción as primary text for semantic comparison; fallback to Requerimiento
        text_for_comparison = desc if desc else nombre

        id_val = cell(row, col_id)
        try:
            req_id = int(float(id_val)) if id_val else req_id
        except (ValueError, TypeError):
            pass

        reqs.append(GoldRequirement(
            id=req_id,
            categoria=_normalize_category(cat_raw),
            tipo=_normalize_type(cell(row, col_tipo)),
            seccion=cell(row, col_sec),
            nombre=nombre,
            descripcion=text_for_comparison,
            documento_formato=cell(row, col_doc),
            categoria_raw=cat_raw,
            tipo_raw=cell(row, col_tipo),
        ))
        req_id += 1

    return reqs


def _parse_indicators_sheet(ws) -> list[GoldIndicator]:
    indicators: list[GoldIndicator] = []
    rows = list(ws.iter_rows(values_only=True))

    for row_idx, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        # Look for rows where first non-empty cell looks like an indicator name
        # and a later cell has a threshold pattern (≥, ≤, >=, <=)
        non_empty = [c for c in cells if c]
        if not non_empty:
            continue

        # Skip header-like rows
        first = non_empty[0].lower()
        if any(skip in first for skip in ["indicador", "tabla", "fórmula", "formula", "ninguno"]):
            continue

        # check if any cell has a threshold pattern
        has_threshold = any(re.search(r"[≥≤<>]=?\s*\d", c) for c in cells)
        if not has_threshold:
            continue

        # first non-empty cell = indicator name
        nombre = non_empty[0]
        # find umbral cell
        umbral_raw = ""
        for c in cells[1:]:
            if re.search(r"[≥≤<>]=?\s*[\d,\.]", c):
                umbral_raw = c
                break

        valor, cond = _parse_threshold(umbral_raw)
        # formula: second non-empty cell if not the threshold
        formula_cells = [c for c in non_empty[1:] if c != umbral_raw and not re.search(r"[≥≤<>]", c)]
        formula = formula_cells[0] if formula_cells else ""

        indicators.append(GoldIndicator(
            nombre=nombre,
            formula=formula,
            umbral_raw=umbral_raw,
            umbral_valor=valor,
            umbral_condicion=cond,
        ))

    return indicators


def _parse_experience_sheet(ws) -> list[GoldExperienceSegment]:
    segments: list[GoldExperienceSegment] = []
    rows = list(ws.iter_rows(values_only=True))

    header_idx = _find_header_row(ws, ["segmento", "nombre", "smmlv", "experiencia"])
    if header_idx is None:
        # try to extract any row with a number + text as a segment
        for row in rows:
            non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if len(non_empty) >= 2:
                # check if first cell is a number (segment id)
                try:
                    float(non_empty[0])
                    nombre = non_empty[1] if len(non_empty) > 1 else ""
                    smmlv = None
                    for c in non_empty[2:]:
                        try:
                            smmlv = float(str(c).replace(",", "."))
                            break
                        except (ValueError, TypeError):
                            continue
                    if nombre:
                        segments.append(GoldExperienceSegment(nombre=nombre, smmlv_minimos=smmlv))
                except (ValueError, TypeError):
                    continue
        return segments

    header = rows[header_idx - 1]
    col_nombre = _col_index(header, ["nombre", "segmento", "descripción", "descripcion"])
    col_smmlv = _col_index(header, ["smmlv", "valor", "mínimo", "minimo"])
    col_desc = _col_index(header, ["tecnologías", "tecnologias", "descripción", "descripcion", "detalle"])

    def cell(row, idx):
        if idx < 0 or idx >= len(row):
            return ""
        return str(row[idx]).strip() if row[idx] is not None else ""

    for row in rows[header_idx:]:
        if all(c is None for c in row):
            continue
        nombre = cell(row, col_nombre)
        if not nombre:
            continue
        smmlv = None
        smmlv_raw = cell(row, col_smmlv)
        try:
            smmlv = float(str(smmlv_raw).replace(",", ".").replace(".", "", smmlv_raw.count(".") - 1))
        except (ValueError, TypeError):
            pass
        desc = cell(row, col_desc) if col_desc >= 0 else ""
        segments.append(GoldExperienceSegment(nombre=nombre, smmlv_minimos=smmlv, descripcion=desc))

    return segments


def parse_gold_standard(excel_path: Path, pdf_name: str) -> GoldStandard:
    """Read a Gold Standard Excel file and return a normalized GoldStandard object."""
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    gs = GoldStandard(pdf_name=pdf_name)

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Requirements
    for candidate in ["requerimientos", "requisitos", "requirements"]:
        if candidate in sheet_names_lower:
            gs.requirements = _parse_requirements_sheet(wb[sheet_names_lower[candidate]])
            break

    # Indicators
    for candidate in ["indicadores", "indicators", "financieros"]:
        if candidate in sheet_names_lower:
            gs.indicators = _parse_indicators_sheet(wb[sheet_names_lower[candidate]])
            break

    # Experience
    for candidate in ["experiencia", "experience"]:
        if candidate in sheet_names_lower:
            gs.experience = _parse_experience_sheet(wb[sheet_names_lower[candidate]])
            break

    wb.close()
    return gs
