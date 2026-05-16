"""
Generate audit reports: INFORME_AUDITORIA_{fecha}.md + resultados_auditoria_{fecha}.xlsx
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from audit.comparator import ComparisonResult, IndicatorMatch

_GREEN = "C6EFCE"
_RED = "FFC7CE"
_YELLOW = "FFEB9C"
_BLUE = "BDD7EE"
_GREY = "D9D9D9"


def _pct(v: float) -> str:
    return f"{v * 100:.1f}%"


def _fmt_time(s: float) -> str:
    if s < 60:
        return f"{s:.0f}s"
    return f"{s / 60:.1f}min"


def _ind_match_summary(matches: list[IndicatorMatch]) -> str:
    if not matches:
        return "N/A"
    matched = sum(1 for m in matches if m.matched)
    total = len(matches)
    gaps = [m.gold_name for m in matches if not m.matched]
    s = f"{matched}/{total} matched"
    if gaps:
        s += f" — gaps: {', '.join(gaps)}"
    return s


def generate_markdown(results: list[ComparisonResult], output_dir: Path) -> Path:
    fecha = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = output_dir / f"INFORME_AUDITORIA_{fecha}.md"

    lines: list[str] = []
    lines.append(f"# Auditoría tendermod vs Gold Standard — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append("## Resumen ejecutivo")
    lines.append("")

    if any(r.semantic_used for r in results):
        lines.append("| PDF | Gold Reqs | TM Reqs | Recall | Precision | F1 | Ind Gold | Ind TM | Tiempo |")
        lines.append("|-----|-----------|---------|--------|-----------|----|----------|--------|--------|")
        for r in results:
            lines.append(
                f"| {r.pdf_name} "
                f"| {r.counts.get('gold_total_reqs', '?')} "
                f"| {r.counts.get('tendermod_total_reqs', '?')} "
                f"| {_pct(r.recall)} "
                f"| {_pct(r.precision)} "
                f"| {_pct(r.f1)} "
                f"| {r.counts.get('gold_indicators_count', '?')} "
                f"| {r.counts.get('tendermod_indicators_count', '?')} "
                f"| {_fmt_time(r.time_total_extraction)} |"
            )
    else:
        lines.append("| PDF | Gold Reqs | TM Reqs | Ind Gold | Ind TM | Tiempo |")
        lines.append("|-----|-----------|---------|----------|--------|--------|")
        for r in results:
            lines.append(
                f"| {r.pdf_name} "
                f"| {r.counts.get('gold_total_reqs', '?')} "
                f"| {r.counts.get('tendermod_total_reqs', '?')} "
                f"| {r.counts.get('gold_indicators_count', '?')} "
                f"| {r.counts.get('tendermod_indicators_count', '?')} "
                f"| {_fmt_time(r.time_total_extraction)} |"
            )

    lines.append("")
    lines.append("## Desglose por categoría (Gold vs Tendermod)")
    lines.append("")

    all_cats = set()
    for r in results:
        all_cats.update(r.counts.get("gold_by_category", {}).keys())
        all_cats.update(r.counts.get("tendermod_by_category", {}).keys())

    for r in results:
        lines.append(f"### {r.pdf_name}")
        lines.append("")
        lines.append("| Categoría | Gold | TM | Delta |")
        lines.append("|-----------|------|----|-------|")
        cats = set(r.counts.get("gold_by_category", {}).keys()) | set(r.counts.get("tendermod_by_category", {}).keys())
        for cat in sorted(cats):
            g = r.counts.get("gold_by_category", {}).get(cat, 0)
            t = r.counts.get("tendermod_by_category", {}).get(cat, 0)
            delta = t - g
            sign = "+" if delta > 0 else ""
            lines.append(f"| {cat} | {g} | {t} | {sign}{delta} |")
        lines.append("")

    lines.append("## Indicadores por PDF")
    lines.append("")
    lines.append("| PDF | Indicadores Gold | Indicadores TM | Detalle |")
    lines.append("|-----|-----------------|----------------|---------|")
    for r in results:
        lines.append(
            f"| {r.pdf_name} "
            f"| {r.counts.get('gold_indicators_count', '?')} "
            f"| {r.counts.get('tendermod_indicators_count', '?')} "
            f"| {_ind_match_summary(r.indicator_matches)} |"
        )

    lines.append("")
    lines.append("## Hallazgos detallados por PDF")
    lines.append("")

    for r in results:
        lines.append(f"### {r.pdf_name}")
        lines.append("")

        if r.semantic_used:
            lines.append(f"**Recall:** {_pct(r.recall)} | **Precision:** {_pct(r.precision)} | **F1:** {_pct(r.f1)}")
            lines.append("")

            if r.gold_unmatched:
                lines.append(f"**Requisitos Gold NO capturados por tendermod ({len(r.gold_unmatched)}):**")
                for req in r.gold_unmatched:
                    cat = req.get("categoria", "?")
                    desc = req.get("descripcion", req.get("nombre", ""))[:150]
                    sec = req.get("seccion", "")
                    lines.append(f"- `[{cat}]` {sec} — {desc}")
                lines.append("")

            if r.tm_unmatched:
                lines.append(f"**Requisitos TM sin match en Gold — posible ruido ({len(r.tm_unmatched)}):**")
                for req in r.tm_unmatched[:20]:
                    cat = req.get("categoria", "?")
                    desc = req.get("descripcion", "")[:120]
                    lines.append(f"- `[{cat}]` {desc}")
                if len(r.tm_unmatched) > 20:
                    lines.append(f"  _... y {len(r.tm_unmatched) - 20} más_")
                lines.append("")

        if r.indicator_matches:
            lines.append("**Indicadores financieros:**")
            for m in r.indicator_matches:
                icon = "✓" if m.matched else "✗"
                th_icon = ""
                if m.threshold_ok is True:
                    th_icon = " (umbral ✓)"
                elif m.threshold_ok is False:
                    th_icon = " (umbral ✗)"
                lines.append(
                    f"- {icon} Gold: `{m.gold_name}` {m.gold_threshold_raw}"
                    + (f" → TM: `{m.tm_name}` = {m.tm_value}{th_icon} (score {m.name_score:.2f})" if m.matched else " → NO encontrado en TM")
                )
            lines.append("")

        exp = r.experience_summary
        if exp:
            lines.append("**Experiencia:**")
            lines.append(f"- Modo TM: `{exp.get('modo', '?')}`")
            lines.append(f"- Códigos UNSPSC TM: {exp.get('codigos_requeridos', exp.get('codigos', []))}")
            lines.append(f"- Valor mínimo TM: {exp.get('valor', '?')}")
            lines.append(f"- Contratos requeridos TM: {exp.get('cantidad_contratos', '?')}")
            gold_segs = exp.get("gold_segments", [])
            if gold_segs:
                lines.append(f"- Segmentos Gold: {', '.join(gold_segs)}")
            lines.append("")

    lines.append("## Patrones de error detectados")
    lines.append("")
    lines.append("_Completar manualmente tras revisar los hallazgos anteriores._")
    lines.append("")
    lines.append("## Recomendaciones de mejora")
    lines.append("")
    lines.append("_Completar manualmente._")
    lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    return path


def generate_excel(results: list[ComparisonResult], output_dir: Path) -> Path:
    fecha = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = output_dir / f"resultados_auditoria_{fecha}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for r in results:
        sheet_name = r.pdf_name[:30].replace("/", "-")
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "Gold_ID", "Gold_Categoria", "Gold_Tipo", "Gold_Seccion",
            "Gold_Descripcion", "Match_Score", "TM_Categoria", "TM_Tipo",
            "TM_Descripcion", "Status",
        ]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor=_BLUE)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)

        # MATCHED rows
        for pair in r.matched_pairs:
            row = [
                pair.get("gold_id", ""),
                pair.get("gold_cat", ""),
                pair.get("gold_tipo", ""),
                pair.get("gold_seccion", pair.get("seccion", "")),
                pair.get("gold_desc", ""),
                pair.get("score", ""),
                pair.get("tm_cat", ""),
                pair.get("tm_tipo", ""),
                pair.get("tm_desc", ""),
                "MATCHED",
            ]
            ws.append(row)
            fill = PatternFill("solid", fgColor=_GREEN)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = fill

        # GOLD_ONLY rows
        for req in r.gold_unmatched:
            row = [
                req.get("id", ""),
                req.get("categoria", ""),
                req.get("tipo", ""),
                req.get("seccion", ""),
                req.get("descripcion", req.get("nombre", ""))[:300],
                "", "", "", "",
                "GOLD_ONLY",
            ]
            ws.append(row)
            fill = PatternFill("solid", fgColor=_RED)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = fill

        # TM_ONLY rows
        for req in r.tm_unmatched:
            row = [
                "",
                "",
                "",
                "",
                "",
                "",
                req.get("categoria", ""),
                req.get("tipo", ""),
                req.get("descripcion", "")[:300],
                "TM_ONLY",
            ]
            ws.append(row)
            fill = PatternFill("solid", fgColor=_YELLOW)
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=ws.max_row, column=col_idx).fill = fill

        # Column widths
        col_widths = [8, 14, 16, 12, 60, 10, 14, 16, 60, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A2"

    # Summary sheet
    ws_sum = wb.create_sheet(title="Resumen", index=0)
    sum_headers = ["PDF", "Gold Reqs", "TM Reqs", "Recall", "Precision", "F1",
                   "Ind Gold", "Ind TM", "Tiempo Total", "Errores"]
    ws_sum.append(sum_headers)
    for col_idx, _ in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=_BLUE)

    for r in results:
        ws_sum.append([
            r.pdf_name,
            r.counts.get("gold_total_reqs", ""),
            r.counts.get("tendermod_total_reqs", ""),
            _pct(r.recall) if r.semantic_used else "N/A",
            _pct(r.precision) if r.semantic_used else "N/A",
            _pct(r.f1) if r.semantic_used else "N/A",
            r.counts.get("gold_indicators_count", ""),
            r.counts.get("tendermod_indicators_count", ""),
            _fmt_time(r.time_total_extraction),
            len(getattr(r, "_errors", [])),
        ])

    for i, w in enumerate([50, 10, 10, 10, 10, 10, 10, 10, 12, 10], 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    wb.save(str(path))
    return path


def generate_report(results: list[ComparisonResult], output_dir: Path) -> tuple[Path, Path]:
    """Generate both markdown and Excel reports. Returns (md_path, xlsx_path)."""
    output_dir.mkdir(parents=True, exist_ok=True)
    md_path = generate_markdown(results, output_dir)
    xlsx_path = generate_excel(results, output_dir)
    return md_path, xlsx_path
