import glob
import json
import logging
import os

from django.conf import settings
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, FileResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.http import require_POST, require_GET

from apps.core.models import AnalysisSession, AnalysisResult, SystemConfig
from .forms import PDFUploadForm, ExperienceEditForm, IndicatorsEditForm
from .tasks import (
    ingest_pdf_task,
    extract_experience_task,
    extract_indicators_task,
    extract_general_info_task,
    extract_general_requirements_task,
    extract_team_profiles_task,
    evaluate_experience_task,
    evaluate_indicators_task,
    evaluate_team_profiles_task,
    generate_conclusion_task,
    finalize_auto_flow_task,
    mark_auto_error_task,
    quick_evaluate_experience_task,
    quick_evaluate_indicators_task,
)

# Lista canonica de pasos del flujo automatico: (nombre_en_timing, label_UI, peso).
# El nombre debe coincidir EXACTO con el usado en _record_timing de cada tarea.
AUTO_FLOW_STEPS = [
    ('Ingesta del PDF', 'Ingesta del PDF en base vectorial', 10),
    ('Extraccion de experiencia RUP', 'Extraccion de requisitos de experiencia', 10),
    ('Extraccion de indicadores financieros', 'Extraccion de indicadores financieros', 10),
    ('Extraccion de informacion general', 'Extraccion de informacion general del proceso', 6),
    ('Extraccion de requisitos habilitantes', 'Extraccion de requisitos habilitantes', 10),
    ('Extraccion de perfiles de equipo de trabajo', 'Extraccion de perfiles de equipo de trabajo', 10),
    ('Evaluacion de cumplimiento de experiencia', 'Evaluacion de cumplimiento de experiencia', 11),
    ('Evaluacion de indicadores financieros', 'Evaluacion de indicadores financieros', 11),
    ('Evaluacion de equipo de trabajo', 'Evaluacion de equipo de trabajo', 12),
    ('Generacion de conclusion ejecutiva', 'Generacion de conclusion ejecutiva', 10),
]

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Vista: lista de sesiones
# ---------------------------------------------------------------------------

def analysis_list(request):
    sessions = AnalysisSession.objects.select_related('result').all()
    return render(request, 'analysis/list.html', {'sessions': sessions})


# ---------------------------------------------------------------------------
# Vista: nueva sesion (subida de PDF)
# ---------------------------------------------------------------------------

def analysis_new(request):
    if request.method == 'POST':
        form = PDFUploadForm(request.POST, request.FILES)
        if form.is_valid():
            pdf_file = form.cleaned_data['pdf_file']
            pdf_filename = pdf_file.name

            # Limpiar PDFs anteriores en data/
            data_dir = settings.TENDERMOD_DATA_DIR
            data_dir.mkdir(parents=True, exist_ok=True)
            for old_pdf in glob.glob(str(data_dir / '*.pdf')):
                os.remove(old_pdf)
                logger.info('PDF anterior eliminado: %s', old_pdf)

            # Copiar el nuevo PDF a data/
            dest = data_dir / pdf_filename
            try:
                with open(dest, 'wb') as f:
                    for chunk in pdf_file.chunks():
                        f.write(chunk)
                logger.info('PDF copiado a %s', dest)
            except Exception as exc:
                logger.error('Error copiando PDF: %s', exc)
                messages.error(request, f'Error al guardar el PDF: {exc}')
                return render(request, 'analysis/new.html', {'form': form})

            # Crear sesion
            session = AnalysisSession.objects.create(
                pdf_filename=pdf_filename,
                status='created',
            )

            action = request.POST.get('action', 'manual')
            if action == 'auto':
                # Flujo automatico: encadenar TODAS las tareas en el backend (chain).
                # El navegador deja de orquestar; solo observa el progreso via auto_progress.
                from celery import chain
                sid = session.pk
                session.timing_json = '[]'  # reset de corridas previas
                session.status = 'auto_running'
                session.auto_flow_active = True
                session.save(update_fields=['timing_json', 'status', 'auto_flow_active', 'updated_at'])
                flow = chain(
                    ingest_pdf_task.si(sid),
                    extract_experience_task.si(sid),
                    extract_indicators_task.si(sid),
                    extract_general_info_task.si(sid),
                    extract_general_requirements_task.si(sid),
                    extract_team_profiles_task.si(sid),
                    evaluate_experience_task.si(sid, None, None, False),
                    evaluate_indicators_task.si(sid, None, False),
                    evaluate_team_profiles_task.si(sid),
                    generate_conclusion_task.si(sid),
                    finalize_auto_flow_task.si(sid),
                )
                async_result = flow.apply_async(link_error=mark_auto_error_task.s(sid))
                session.celery_task_id = async_result.id
                session.save(update_fields=['celery_task_id', 'updated_at'])
                logger.info('Sesion %s creada, chain automatico %s lanzado', sid, async_result.id)
                return redirect('analysis:analysis_auto', pk=session.pk)

            # Flujo manual: solo ingesta, el usuario avanza paso a paso
            task = ingest_pdf_task.delay(session.pk)
            session.celery_task_id = task.id
            session.save(update_fields=['celery_task_id', 'updated_at'])
            logger.info('Sesion %s creada, tarea de ingesta %s lanzada', session.pk, task.id)
            return redirect('analysis:step1', pk=session.pk)
        else:
            messages.error(request, 'Formulario invalido: ' + str(form.errors))
    else:
        form = PDFUploadForm()

    return render(request, 'analysis/new.html', {'form': form})


# ---------------------------------------------------------------------------
# Vista: Paso 1 — Extraccion de requisitos
# ---------------------------------------------------------------------------

def analysis_step1(request, pk):
    session = get_object_or_404(AnalysisSession, pk=pk)
    context = {
        'session': session,
        'has_experience': bool(session.experience_requirements_json),
        'has_indicators': bool(session.indicators_requirements_json),
        'has_general_info': bool(session.general_info_text),
        'has_general_requirements': bool(session.general_requirements_json),
        'has_team_profiles': bool(session.team_profiles_json),
    }
    return render(request, 'analysis/step1.html', context)


@require_POST
def analysis_extract(request, pk):
    """
    AJAX endpoint para lanzar tareas de extraccion desde el Paso 1.
    Body JSON: {action: "experience" | "indicators" | "general_info"}
    Responde con {task_id: "..."}
    """
    session = get_object_or_404(AnalysisSession, pk=pk)

    try:
        body = json.loads(request.body)
        action = body.get('action')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Body JSON invalido'}, status=400)

    if action == 'experience':
        task = extract_experience_task.delay(session.pk)
    elif action == 'indicators':
        task = extract_indicators_task.delay(session.pk)
    elif action == 'general_info':
        task = extract_general_info_task.delay(session.pk)
    elif action == 'general_requirements':
        task = extract_general_requirements_task.delay(session.pk)
    elif action == 'team_profiles':
        task = extract_team_profiles_task.delay(session.pk)
    else:
        return JsonResponse({'error': f'Accion desconocida: {action}'}, status=400)

    session.celery_task_id = task.id
    session.save(update_fields=['celery_task_id', 'updated_at'])

    logger.info('Tarea %s lanzada para sesion %s (action=%s)', task.id, pk, action)
    return JsonResponse({'task_id': task.id})


# ---------------------------------------------------------------------------
# Vista: Paso 2 — Validacion humana y evaluacion
# ---------------------------------------------------------------------------

def analysis_step2(request, pk):
    session = get_object_or_404(AnalysisSession, pk=pk)

    # Pre-poblar formularios desde JSON guardado en la sesion
    exp_initial = {}
    if session.experience_requirements_json:
        try:
            from tendermod.evaluation.schemas import ExperienceResponse
            exp = ExperienceResponse.model_validate_json(session.experience_requirements_json)
            exp_initial = {
                'listado_codigos': ', '.join(exp.listado_codigos),
                'cantidad_codigos': exp.cantidad_codigos,
                'objeto': exp.objeto,
                'cantidad_contratos': exp.cantidad_contratos,
                'valor': exp.valor,
                'pagina': exp.pagina,
                'seccion': exp.seccion,
                'regla_codigos': exp.regla_codigos,
                'objeto_exige_relevancia': exp.objeto_exige_relevancia,
            }
        except Exception as exc:
            logger.error('Error parseando experience_requirements_json: %s', exc)
            messages.warning(request, 'No se pudo cargar la extraccion de experiencia. Revise los datos.')

    ind_initial = {}
    indicators_list = []
    if session.indicators_requirements_json:
        try:
            from tendermod.evaluation.schemas import MultipleIndicatorResponse
            ind = MultipleIndicatorResponse.model_validate_json(session.indicators_requirements_json)
            indicators_list = [{'indicador': i.indicador, 'valor': str(i.valor)} for i in ind.answer]
            ind_initial = {'indicators_json': json.dumps(indicators_list)}
        except Exception as exc:
            logger.error('Error parseando indicators_requirements_json: %s', exc)
            messages.warning(request, 'No se pudo cargar la extraccion de indicadores.')

    if request.method == 'POST':
        # Handle normal form submission (non-AJAX evaluation triggers)
        # AJAX evaluations are handled by analysis_evaluate endpoint
        pass

    general_requirements = []
    if session.general_requirements_json:
        try:
            from tendermod.evaluation.schemas import GeneralRequirementList
            gr = GeneralRequirementList.model_validate_json(session.general_requirements_json)
            general_requirements = gr.requisitos
        except Exception as exc:
            logger.error('Error parseando general_requirements_json en step2: %s', exc)

    exp_form = ExperienceEditForm(initial=exp_initial)
    ind_form = IndicatorsEditForm(initial=ind_initial)

    # Parsear objeto completo ExperienceResponse para acceso en la plantilla
    exp_data = None
    if session.experience_requirements_json:
        try:
            from tendermod.evaluation.schemas import ExperienceResponse
            exp_data = ExperienceResponse.model_validate_json(session.experience_requirements_json)
        except Exception as exc:
            logger.error('Error parseando ExperienceResponse para step2: %s', exc)

    team_profiles = []
    if session.team_profiles_json:
        try:
            from tendermod.evaluation.schemas import ProfileRequirementList
            tp = ProfileRequirementList.model_validate_json(session.team_profiles_json)
            team_profiles = tp.perfiles
        except Exception as exc:
            logger.error('Error parseando team_profiles_json en step2: %s', exc)

    team_compliance = None
    try:
        result_obj = session.result
        if result_obj and result_obj.team_compliance_json:
            from tendermod.evaluation.schemas import TeamProfileComplianceList
            team_compliance = TeamProfileComplianceList.model_validate_json(result_obj.team_compliance_json)
    except AnalysisResult.DoesNotExist:
        pass
    except Exception as exc:
        logger.error('Error parseando team_compliance_json en step2: %s', exc)

    context = {
        'session': session,
        'exp_form': exp_form,
        'ind_form': ind_form,
        'indicators_list': indicators_list,
        'exp_initial_json': json.dumps(exp_initial),
        'system_threshold': SystemConfig.get_solo().threshold_objeto,
        'exp_data': exp_data,
        'general_requirements': general_requirements,
        'req_count': len(general_requirements),
        'team_profiles': team_profiles,
        'team_compliance': team_compliance,
    }
    return render(request, 'analysis/step2.html', context)


@require_POST
def analysis_evaluate(request, pk):
    """
    AJAX endpoint para lanzar evaluaciones desde el Paso 2.
    Body JSON:
      {action: "experience", experience_data: {...}}  — evalua experiencia
      {action: "indicators", indicators_list: [...]}  — evalua indicadores
    """
    session = get_object_or_404(AnalysisSession, pk=pk)

    try:
        body = json.loads(request.body)
        action = body.get('action')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Body JSON invalido'}, status=400)

    if action == 'experience':
        exp_data = body.get('experience_data', {})
        # Normalizar el campo listado_codigos desde string a lista
        if isinstance(exp_data.get('listado_codigos'), str):
            raw = exp_data['listado_codigos']
            exp_data['listado_codigos'] = [c.strip() for c in raw.split(',') if c.strip()]
        # Extraer umbral de experience_data (enviado por buildExperienceData) o del cuerpo
        # y removerlo del dict antes de construir ExperienceResponse
        try:
            umbral = float(exp_data.pop('umbral', body.get('umbral', SystemConfig.get_solo().threshold_objeto)))
        except (TypeError, ValueError):
            umbral = float(SystemConfig.get_solo().threshold_objeto)
        task = evaluate_experience_task.delay(session.pk, exp_data, umbral)

    elif action == 'indicators':
        ind_list = body.get('indicators_list', [])
        task = evaluate_indicators_task.delay(session.pk, ind_list)

    elif action == 'team_profiles':
        task = evaluate_team_profiles_task.delay(session.pk)

    else:
        return JsonResponse({'error': f'Accion desconocida: {action}'}, status=400)

    session.celery_task_id = task.id
    session.save(update_fields=['celery_task_id', 'updated_at'])

    logger.info('Tarea de evaluacion %s lanzada para sesion %s (action=%s)', task.id, pk, action)
    return JsonResponse({'task_id': task.id})


# ---------------------------------------------------------------------------
# Vista: Resultados
# ---------------------------------------------------------------------------

def analysis_results(request, pk):
    session = get_object_or_404(AnalysisSession, pk=pk)

    result = None
    exp_result = None
    ind_result = None

    try:
        result = session.result
    except AnalysisResult.DoesNotExist:
        pass

    if result:
        if result.experience_result_json:
            try:
                from tendermod.evaluation.schemas import ExperienceComplianceResult
                exp_result = ExperienceComplianceResult.model_validate_json(result.experience_result_json)
            except Exception as exc:
                logger.error('Error parseando experience_result_json: %s', exc)

        if result.indicators_result_json:
            try:
                from tendermod.evaluation.schemas import IndicatorComplianceResult
                ind_result = IndicatorComplianceResult.model_validate_json(result.indicators_result_json)
            except Exception as exc:
                logger.error('Error parseando indicators_result_json: %s', exc)

    general_requirements = []
    if session.general_requirements_json:
        try:
            from tendermod.evaluation.schemas import GeneralRequirementList
            gr = GeneralRequirementList.model_validate_json(session.general_requirements_json)
            general_requirements = gr.requisitos
        except Exception as exc:
            logger.error('Error parseando general_requirements_json en results: %s', exc)

    team_compliance = None
    if result and result.team_compliance_json:
        try:
            from tendermod.evaluation.schemas import TeamProfileComplianceList
            team_compliance = TeamProfileComplianceList.model_validate_json(result.team_compliance_json)
        except Exception as exc:
            logger.error('Error parseando team_compliance_json en results: %s', exc)

    conclusion = None
    if result and result.conclusion_json:
        try:
            from tendermod.evaluation.schemas import EvaluacionConclusionResult
            conclusion = EvaluacionConclusionResult.model_validate_json(result.conclusion_json)
        except Exception as exc:
            logger.error('Error parseando conclusion_json en results: %s', exc)

    context = {
        'session': session,
        'result': result,
        'general_requirements': general_requirements,
        'exp_result': exp_result,
        'ind_result': ind_result,
        'team_compliance': team_compliance,
        'conclusion': conclusion,
    }
    return render(request, 'analysis/results.html', context)


# ---------------------------------------------------------------------------
# Requisitos Generales: guardar estados del checklist
# ---------------------------------------------------------------------------

@require_POST
def analysis_checklist_save(request, pk):
    """
    AJAX: recibe lista de {id, estado} y persiste los estados en general_requirements_json.
    Body: {updates: [{id: 1, estado: "CUMPLE"}, ...]}
    """
    session = get_object_or_404(AnalysisSession, pk=pk)
    try:
        body = json.loads(request.body)
        updates = {
            item['id']: {'estado': item.get('estado'), 'nota': item.get('nota', '')}
            for item in body.get('updates', [])
        }
    except (json.JSONDecodeError, KeyError, TypeError):
        return JsonResponse({'error': 'Body JSON invalido'}, status=400)

    if not session.general_requirements_json:
        return JsonResponse({'error': 'No hay requisitos cargados'}, status=400)

    try:
        from tendermod.evaluation.schemas import GeneralRequirementList
        req_list = GeneralRequirementList.model_validate_json(session.general_requirements_json)
        for req in req_list.requisitos:
            if req.id in updates:
                upd = updates[req.id]
                if upd.get('estado'):
                    req.estado = upd['estado']
                req.nota = upd.get('nota', req.nota)
        session.general_requirements_json = req_list.model_dump_json()
        session.save(update_fields=['general_requirements_json', 'updated_at'])
    except Exception as exc:
        logger.error('Error guardando checklist para sesion %s: %s', pk, exc)
        return JsonResponse({'error': str(exc)}, status=500)

    return JsonResponse({'status': 'ok'})


@require_POST
def analysis_pliego_qa(request, pk):
    """
    Responde una pregunta sobre el pliego via RAG (sincrono).
    Si add_as_requirement=True, agrega la respuesta como GeneralRequirement (origen="QA").
    Body: {question: "...", add_as_requirement: bool, categoria: "OTRO"}
    """
    session = get_object_or_404(AnalysisSession, pk=pk)
    try:
        body = json.loads(request.body)
        question = body.get('question', '').strip()
        add_as_requirement = body.get('add_as_requirement', False)
        categoria = body.get('categoria', 'OTRO')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Body JSON invalido'}, status=400)

    if not question:
        return JsonResponse({'error': 'Pregunta vacia'}, status=400)

    try:
        from tendermod.evaluation.general_requirements_inference import ask_pliego
        answer = ask_pliego(question, k=8)
    except Exception as exc:
        logger.error('Error en Q&A del pliego para sesion %s: %s', pk, exc)
        return JsonResponse({'error': str(exc)}, status=500)

    if add_as_requirement and answer and 'No se encontr' not in answer:
        try:
            from tendermod.evaluation.schemas import GeneralRequirementList, GeneralRequirement
            req_list = GeneralRequirementList()
            if session.general_requirements_json:
                req_list = GeneralRequirementList.model_validate_json(session.general_requirements_json)
            next_id = max((r.id for r in req_list.requisitos), default=0) + 1
            req_list.requisitos.append(GeneralRequirement(
                id=next_id,
                categoria=categoria,
                descripcion=f"[Q&A] {question[:120]}: {answer[:300]}",
                obligatorio='NO_ESPECIFICADO',
                estado='PENDIENTE',
                origen='QA',
            ))
            session.general_requirements_json = req_list.model_dump_json()
            session.save(update_fields=['general_requirements_json', 'updated_at'])
        except Exception as exc:
            logger.error('Error agregando requisito Q&A para sesion %s: %s', pk, exc)

    return JsonResponse({'answer': answer, 'added': add_as_requirement})


# ---------------------------------------------------------------------------
# Exportacion
# ---------------------------------------------------------------------------

def export_excel(request, pk):
    """
    Exporta los resultados de la sesion como archivo Excel (.xlsx).
    Hoja 1: Indicadores | Hoja 2: Experiencia RUP | Hoja 3: Checklist General | Hoja 4: Equipo de Trabajo
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from io import BytesIO

    def _apply_table_format(ws, table_name):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        if ws.max_row >= 1 and ws.max_column >= 1:
            tab = Table(displayName=table_name, ref=ws.dimensions)
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(tab)

    session = get_object_or_404(AnalysisSession, pk=pk)

    try:
        result = session.result
    except AnalysisResult.DoesNotExist:
        messages.error(request, 'No hay resultados para exportar.')
        return redirect('analysis:results', pk=pk)

    wb = openpyxl.Workbook()

    # ---- Hoja 1: Indicadores ----
    ws_ind = wb.active
    ws_ind.title = 'Indicadores'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E79')

    ind_headers = ['Indicador', 'Valor Empresa', 'Condicion', 'Umbral', 'Cumple', 'Detalle']
    for col_num, header in enumerate(ind_headers, 1):
        cell = ws_ind.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    if result.indicators_result_json:
        try:
            import re as _re
            from tendermod.evaluation.schemas import IndicatorComplianceResult
            ind_result = IndicatorComplianceResult.model_validate_json(result.indicators_result_json)
            cumple_str = 'CUMPLE' if ind_result.cumple else ('NO CUMPLE' if ind_result.cumple is False else 'INDETERMINADO')

            # Parsear secciones numeradas del texto LLM
            raw = ind_result.detalle.strip()
            partes = _re.split(r'(?m)(?=^\d+\. )', raw)
            secciones_numeradas = [p.strip() for p in partes if p.strip() and _re.match(r'^\d+\.', p.strip())]
            # Texto de conclusión: lo que queda tras la última sección numerada
            resto = _re.split(r'(?m)(?=^\*\*Conclusi)', raw)
            conclusion_text = resto[-1].strip() if len(resto) > 1 else ''

            if ind_result.indicadores_detalle:
                for i, det in enumerate(ind_result.indicadores_detalle):
                    cumple_ind = 'SI' if det.cumple else ('NO' if det.cumple is False else '-')
                    val_str = f'{det.valor_empresa:g}' if det.valor_empresa is not None else ''
                    umb_str = f'{det.umbral:g}' if det.umbral is not None else ''
                    detalle_txt = secciones_numeradas[i] if i < len(secciones_numeradas) else ''
                    row_num = ws_ind.max_row + 1
                    ws_ind.append([det.indicador, val_str, det.condicion or '', umb_str, cumple_ind, detalle_txt])
                    cell_cumple = ws_ind.cell(row=row_num, column=5)
                    if det.cumple is True:
                        cell_cumple.fill = PatternFill('solid', fgColor='C6EFCE')
                    elif det.cumple is False:
                        cell_cumple.fill = PatternFill('solid', fgColor='FFC7CE')
            else:
                for nombre in ind_result.indicadores_evaluados:
                    ws_ind.append([nombre, '', '', '', '', ''])

            # Fila de conclusión final
            concl_row = ws_ind.max_row + 1
            ws_ind.append(['CONCLUSION FINAL', '', '', '', cumple_str, conclusion_text])
            concl_fill = (PatternFill('solid', fgColor='C6EFCE') if ind_result.cumple else
                          PatternFill('solid', fgColor='FFC7CE') if ind_result.cumple is False else
                          PatternFill('solid', fgColor='FFEB9C'))
            for c in range(1, 7):
                ws_ind.cell(row=concl_row, column=c).fill = concl_fill

        except Exception:
            ws_ind.append(['Error al parsear resultado de indicadores', '', '', '', '', ''])

    # Auto-width columns
    for col in ws_ind.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws_ind.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    _apply_table_format(ws_ind, "IndTable")

    # ---- Hoja 2: Experiencia (unificada: RUPs + Sub-Requisitos) ----
    ws_exp = wb.create_sheet('Experiencia')

    exp_headers = ['NUMERO RUP', 'Cliente', 'Valor COP', 'Cumple Codigos',
                   'Cumple Valor', 'Cumple Objeto', 'Score Objeto',
                   'Contrato Elegido', 'Cumple Total']
    for col_num, header in enumerate(exp_headers, 1):
        cell = ws_exp.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    def fmt_cop(v):
        try:
            return f'${int(round(float(v))):,}'.replace(',', '.')
        except Exception:
            return v

    rup_end_row = 1

    if result.experience_result_json:
        try:
            from tendermod.evaluation.schemas import ExperienceComplianceResult
            exp_result = ExperienceComplianceResult.model_validate_json(result.experience_result_json)

            # Summary row
            ws_exp.cell(row=2, column=1, value='RESUMEN')
            ws_exp.cell(row=2, column=2, value=f'Codigos requeridos: {", ".join(exp_result.codigos_requeridos)}')
            ws_exp.cell(row=2, column=3, value=fmt_cop(exp_result.valor_requerido_cop) if exp_result.valor_requerido_cop else '')
            ws_exp.cell(row=2, column=9, value='CUMPLE' if exp_result.cumple else 'NO CUMPLE')

            for rup in exp_result.rups_evaluados:
                ws_exp.append([
                    rup.numero_rup,
                    rup.cliente or '',
                    fmt_cop(rup.valor_cop) if rup.valor_cop else '',
                    'SI' if rup.cumple_codigos else 'NO',
                    'SI' if rup.cumple_valor else ('NO' if rup.cumple_valor is False else 'N/A'),
                    'SI' if rup.cumple_objeto else ('NO' if rup.cumple_objeto is False else 'N/A'),
                    f'{rup.score_objeto:.3f}' if rup.score_objeto is not None else 'N/A',
                    (rup.objeto_contrato or '')[:200],
                    'CUMPLE' if rup.cumple_total else 'NO CUMPLE',
                ])

            rup_end_row = ws_exp.max_row

            # Sub-requisitos section (modo MULTI_CONDICION)
            if exp_result.sub_requisitos_resultado:
                sep_row = ws_exp.max_row + 1
                sep_fill = PatternFill('solid', fgColor='2E4057')
                sep_font_style = Font(bold=True, color='FFFFFF')
                ws_exp.cell(row=sep_row, column=1, value='── SUB-REQUISITOS ──')
                for c in range(1, len(exp_headers) + 1):
                    ws_exp.cell(row=sep_row, column=c).fill = sep_fill
                    ws_exp.cell(row=sep_row, column=c).font = sep_font_style

                sub_header_row = sep_row + 1
                sub_headers = ['NUMERO RUP', 'Sub-Requisito', 'Cumple', 'Score', 'Contrato que Cumple']
                for col_num, hdr in enumerate(sub_headers, 1):
                    c = ws_exp.cell(row=sub_header_row, column=col_num, value=hdr)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal='center')

                for sr in exp_result.sub_requisitos_resultado:
                    ws_exp.append([
                        sr.rup_elegido if sr.rup_elegido is not None else '',
                        sr.descripcion,
                        'CUMPLE' if sr.cumple else 'NO CUMPLE',
                        f'{sr.score_objeto:.3f}' if sr.score_objeto is not None else 'N/A',
                        (sr.objeto_contrato or '')[:200],
                    ])
                    for cell in ws_exp[ws_exp.max_row]:
                        cell.alignment = Alignment(wrap_text=True)

        except Exception as exc:
            ws_exp.append([f'Error al parsear resultado de experiencia: {exc}'])

    for col in ws_exp.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws_exp.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Tabla solo sobre el bloque RUPs; sub-requisitos quedan fuera del rango
    from openpyxl.utils import get_column_letter as _gcl
    for row in ws_exp.iter_rows(min_row=2, max_row=max(rup_end_row, 1)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    if rup_end_row > 1:
        rup_ref = f"A1:{_gcl(len(exp_headers))}{rup_end_row}"
        tab_exp = Table(displayName="ExpTable", ref=rup_ref)
        tab_exp.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws_exp.add_table(tab_exp)

    # ---- Hoja: Checklist General ----
    if session.general_requirements_json:
        try:
            from tendermod.evaluation.schemas import GeneralRequirementList
            import urllib.parse as _urlparse
            _pdf_link = _urlparse.quote(session.pdf_filename or '', safe='')
            gr = GeneralRequirementList.model_validate_json(session.general_requirements_json)
            if gr.requisitos:
                ws_cl = wb.create_sheet('Checklist General')

                cl_headers = [
                    '#', 'Categoria', 'Tipo', 'Descripcion', 'Extracto Pliego',
                    'Documento/Formato', 'Obligatorio', 'Seccion', 'Pagina',
                    'Estado', 'Nota', 'Origen', 'Confianza', 'Cita Verificada',
                ]
                for col_num, header in enumerate(cl_headers, 1):
                    cell = ws_cl.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                tipo_fills = {
                    'HABILITANTE-EXPERIENCIA':  PatternFill('solid', fgColor='D6E4BC'),  # verde oliva
                    'HABILITANTE-INDICADORES':  PatternFill('solid', fgColor='D9E1F2'),  # azul lavanda
                    'HABILITANTE':    PatternFill('solid', fgColor='DEEAF1'),  # azul claro
                    'PUNTUABLE':      PatternFill('solid', fgColor='E2EFDA'),  # verde claro
                    'DOCUMENTAL':     PatternFill('solid', fgColor='FFF2CC'),  # amarillo claro
                    'GARANTIA':       PatternFill('solid', fgColor='EAD1DC'),  # rosado claro
                    'CAUSAL_RECHAZO': PatternFill('solid', fgColor='FCE4D6'),  # salmon claro
                    'NO_ESPECIFICADO': PatternFill('solid', fgColor='F2F2F2'), # gris claro
                }

                for req in gr.requisitos:
                    row_num = ws_cl.max_row + 1
                    confidence_pct = f"{req.confidence * 100:.0f}%" if req.confidence is not None else ""
                    citation_label = {True: "Verificada", False: "No encontrada", None: ""}.get(req.citation_verified, "")
                    ws_cl.append([
                        req.id, req.categoria, req.tipo, req.descripcion,
                        (req.extracto_pliego or '')[:600],
                        req.documento_formato, req.obligatorio, req.seccion,
                        req.pagina, req.estado, getattr(req, 'nota', ''), req.origen,
                        confidence_pct, citation_label,
                    ])
                    fill = tipo_fills.get(req.tipo)
                    if fill:
                        for col in range(1, 15):
                            ws_cl.cell(row=row_num, column=col).fill = fill
                    # Categoría EXPERIENCIA tiene color propio (naranja) que sobrescribe el tipo
                    if req.categoria == 'EXPERIENCIA':
                        exp_fill = PatternFill('solid', fgColor='FFE4B5')
                        for col in range(1, 15):
                            ws_cl.cell(row=row_num, column=col).fill = exp_fill
                    # Hipervínculo en columna Pagina (col 9) → abre PDF en la página indicada
                    try:
                        page_num = int(str(req.pagina).strip())
                        pagina_cell = ws_cl.cell(row=row_num, column=9)
                        pagina_cell.value = f'p.{page_num}'
                        pagina_cell.hyperlink = f'{_pdf_link}#page={page_num}'
                        pagina_cell.font = Font(color='0563C1', underline='single')
                    except (ValueError, TypeError):
                        pass

                for col in ws_cl.columns:
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws_cl.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
                _apply_table_format(ws_cl, "CLTable")
        except Exception as exc:
            logger.warning('No se pudo generar hoja Checklist General: %s', exc)

    # ---- Hoja: Equipo de Trabajo ----
    if result.team_compliance_json:
        try:
            from tendermod.evaluation.schemas import TeamProfileComplianceList
            team = TeamProfileComplianceList.model_validate_json(result.team_compliance_json)
            if team.perfiles_evaluados:
                ws_team = wb.create_sheet('Equipo de Trabajo')

                team_headers = [
                    'Rol', 'Requeridos', 'Persona', 'Cargo',
                    'Cumple Persona', 'Justificacion', 'Evidencia', 'Gaps',
                ]
                for col_num, header in enumerate(team_headers, 1):
                    cell = ws_team.cell(row=1, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

                fill_ok  = PatternFill('solid', fgColor='C6EFCE')
                fill_no  = PatternFill('solid', fgColor='FFC7CE')
                fill_rol = PatternFill('solid', fgColor='2E4057')
                font_rol = Font(bold=True, color='FFFFFF')

                for perfil in team.perfiles_evaluados:
                    # Fila separadora por perfil
                    sep_row = ws_team.max_row + 1
                    candidatos_txt = (
                        ', '.join(perfil.personas_que_cumplen)
                        if perfil.personas_que_cumplen else '—'
                    )
                    resultado_perfil = 'CUMPLE' if perfil.cumple else 'NO CUMPLE'
                    ws_team.cell(row=sep_row, column=1, value=perfil.rol)
                    ws_team.cell(row=sep_row, column=2, value=perfil.cantidad_requerida)
                    ws_team.cell(row=sep_row, column=5, value=resultado_perfil)
                    ws_team.cell(row=sep_row, column=6, value=f'Candidatos aptos: {candidatos_txt}')
                    for c in range(1, len(team_headers) + 1):
                        ws_team.cell(row=sep_row, column=c).fill = fill_rol
                        ws_team.cell(row=sep_row, column=c).font = font_rol

                    # Una fila por persona evaluada
                    for persona in perfil.personas_evaluadas:
                        row_num = ws_team.max_row + 1
                        ws_team.append([
                            perfil.rol,
                            perfil.cantidad_requerida,
                            persona.persona,
                            persona.cargo,
                            'SI' if persona.cumple else 'NO',
                            persona.justificacion,
                            '; '.join(persona.evidencia) if persona.evidencia else '',
                            '; '.join(persona.gaps) if persona.gaps else '',
                        ])
                        row_fill = fill_ok if persona.cumple else fill_no
                        for c in range(1, len(team_headers) + 1):
                            ws_team.cell(row=row_num, column=c).fill = row_fill

                # Fila de conclusión final
                concl_row = ws_team.max_row + 1
                cumple_equipo_str = 'CUMPLE' if team.cumple_equipo else 'NO CUMPLE'
                ws_team.append(['CONCLUSION FINAL', '', '', '', cumple_equipo_str, '', '', ''])
                concl_fill = fill_ok if team.cumple_equipo else fill_no
                concl_font = Font(bold=True)
                for c in range(1, len(team_headers) + 1):
                    ws_team.cell(row=concl_row, column=c).fill = concl_fill
                    ws_team.cell(row=concl_row, column=c).font = concl_font

                for col in ws_team.columns:
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws_team.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
                for row in ws_team.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')

        except Exception as exc:
            logger.warning('No se pudo generar hoja Equipo de Trabajo: %s', exc)

    # ---- Hoja: Conclusion Ejecutiva ----
    if result.conclusion_json:
        try:
            from tendermod.evaluation.schemas import EvaluacionConclusionResult
            concl = EvaluacionConclusionResult.model_validate_json(result.conclusion_json)
            ws_concl = wb.create_sheet('Conclusion Ejecutiva')

            # Título
            ws_concl.cell(row=1, column=1, value='CONCLUSION EJECUTIVA').font = Font(bold=True, size=14, color='FFFFFF')
            ws_concl.cell(row=1, column=1).fill = PatternFill('solid', fgColor='1F4E79')
            ws_concl.merge_cells('A1:D1')

            # Veredicto general
            ws_concl.cell(row=3, column=1, value='VEREDICTO GENERAL').font = Font(bold=True, color='FFFFFF')
            ws_concl.cell(row=3, column=1).fill = PatternFill('solid', fgColor='2E4057')
            ws_concl.merge_cells('A3:D3')
            ws_concl.cell(row=4, column=1, value=concl.veredicto_general)
            ws_concl.cell(row=4, column=1).alignment = Alignment(wrap_text=True)
            ws_concl.merge_cells('A4:D4')
            ws_concl.row_dimensions[4].height = 90

            # RUPs recomendados
            if concl.rups_recomendados:
                r = ws_concl.max_row + 2
                ws_concl.cell(row=r, column=1, value='CONTRATOS RUP A PRESENTAR').font = Font(bold=True, color='FFFFFF')
                ws_concl.cell(row=r, column=1).fill = PatternFill('solid', fgColor='375623')
                ws_concl.merge_cells(f'A{r}:D{r}')
                r += 1
                for hdr, col in [('N° RUP', 1), ('Cliente', 2), ('Valor COP', 3), ('Relevancia', 4)]:
                    c = ws_concl.cell(row=r, column=col, value=hdr)
                    c.font = header_font
                    c.fill = header_fill
                r += 1
                for rup in concl.rups_recomendados:
                    val_str = f'${int(round(rup.valor_cop)):,}'.replace(',', '.') if rup.valor_cop else ''
                    ws_concl.append([str(rup.numero_rup), rup.cliente or '', val_str, rup.relevancia])
                    for col in range(1, 5):
                        ws_concl.cell(row=ws_concl.max_row, column=col).alignment = Alignment(wrap_text=True)

            # Equipo recomendado
            if concl.personas_recomendadas:
                r = ws_concl.max_row + 2
                ws_concl.cell(row=r, column=1, value='EQUIPO RECOMENDADO').font = Font(bold=True, color='FFFFFF')
                ws_concl.cell(row=r, column=1).fill = PatternFill('solid', fgColor='7B3F00')
                ws_concl.merge_cells(f'A{r}:D{r}')
                r += 1
                for hdr, col in [('Perfil Requerido', 1), ('Personas Aptas', 2)]:
                    c = ws_concl.cell(row=r, column=col, value=hdr)
                    c.font = header_font
                    c.fill = header_fill
                for persona_rec in concl.personas_recomendadas:
                    ws_concl.append([persona_rec.rol, ', '.join(persona_rec.personas)])
                    for col in range(1, 3):
                        ws_concl.cell(row=ws_concl.max_row, column=col).alignment = Alignment(wrap_text=True)

            # Brechas
            if concl.brechas:
                r = ws_concl.max_row + 2
                ws_concl.cell(row=r, column=1, value='BRECHAS DETECTADAS').font = Font(bold=True, color='FFFFFF')
                ws_concl.cell(row=r, column=1).fill = PatternFill('solid', fgColor='9C1A1A')
                ws_concl.merge_cells(f'A{r}:D{r}')
                for i, brecha in enumerate(concl.brechas, 1):
                    ws_concl.append([f'{i}. {brecha}', '', '', ''])
                    ws_concl.cell(row=ws_concl.max_row, column=1).alignment = Alignment(wrap_text=True)
                    ws_concl.merge_cells(f'A{ws_concl.max_row}:D{ws_concl.max_row}')

            # Recomendaciones
            if concl.recomendaciones:
                r = ws_concl.max_row + 2
                ws_concl.cell(row=r, column=1, value='RECOMENDACIONES').font = Font(bold=True, color='FFFFFF')
                ws_concl.cell(row=r, column=1).fill = PatternFill('solid', fgColor='1F4E79')
                ws_concl.merge_cells(f'A{r}:D{r}')
                for i, rec in enumerate(concl.recomendaciones, 1):
                    ws_concl.append([f'{i}. {rec}', '', '', ''])
                    ws_concl.cell(row=ws_concl.max_row, column=1).alignment = Alignment(wrap_text=True)
                    ws_concl.merge_cells(f'A{ws_concl.max_row}:D{ws_concl.max_row}')

            ws_concl.column_dimensions['A'].width = 60
            ws_concl.column_dimensions['B'].width = 35
            ws_concl.column_dimensions['C'].width = 20
            ws_concl.column_dimensions['D'].width = 50
        except Exception as exc:
            logger.warning('No se pudo generar hoja Conclusion Ejecutiva: %s', exc)

    # ---- Hoja: Tiempos de Ejecucion ----
    try:
        import json as _json
        timing_entries = _json.loads(session.timing_json or '[]')
        if timing_entries:
            ws_tim = wb.create_sheet('Tiempos de Ejecucion')
            tim_headers = ['Paso', 'Tarea Celery', 'Duracion (s)', 'Estado']
            for col_num, header in enumerate(tim_headers, 1):
                cell = ws_tim.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            fill_ok  = PatternFill('solid', fgColor='C6EFCE')
            fill_err = PatternFill('solid', fgColor='FFC7CE')
            total_s = 0.0
            for entry in timing_entries:
                dur = entry.get('duracion_s', 0)
                total_s += dur
                estado = entry.get('estado', 'ok')
                row_num = ws_tim.max_row + 1
                ws_tim.append([
                    entry.get('paso', ''),
                    entry.get('tarea', ''),
                    dur,
                    estado.upper(),
                ])
                row_fill = fill_ok if estado == 'ok' else fill_err
                for c in range(1, 5):
                    ws_tim.cell(row=row_num, column=c).fill = row_fill
            # Fila de total
            total_row = ws_tim.max_row + 1
            ws_tim.cell(row=total_row, column=1, value='TOTAL')
            ws_tim.cell(row=total_row, column=3, value=round(total_s, 1))
            for c in range(1, 5):
                ws_tim.cell(row=total_row, column=c).font = Font(bold=True)
            for col in ws_tim.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws_tim.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
            _apply_table_format(ws_tim, "TimTable")
    except Exception as exc:
        logger.warning('No se pudo generar hoja Tiempos de Ejecucion: %s', exc)

    # ---- Generar respuesta ----
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'tendermod_resultados_{pk}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_text(request, pk):
    """
    Exporta los resultados como texto plano con resumen ejecutivo.
    """
    session = get_object_or_404(AnalysisSession, pk=pk)

    try:
        result = session.result
    except AnalysisResult.DoesNotExist:
        messages.error(request, 'No hay resultados para exportar.')
        return redirect('analysis:results', pk=pk)

    lines = []
    lines.append('=' * 70)
    lines.append('TENDERMOD — RESUMEN DE EVALUACION DE CUMPLIMIENTO')
    lines.append('=' * 70)
    lines.append(f'Sesion ID   : {session.pk}')
    lines.append(f'PDF         : {session.pdf_filename}')
    lines.append(f'Fecha       : {session.created_at.strftime("%Y-%m-%d %H:%M")}')
    lines.append(f'Estado      : {session.get_status_display()}')
    lines.append('')

    # Resultado general
    if result.cumple_final is not None:
        veredicto = 'CUMPLE' if result.cumple_final else 'NO CUMPLE'
        lines.append(f'VEREDICTO FINAL: {veredicto}')
    lines.append('')

    # Indicadores
    lines.append('-' * 70)
    lines.append('INDICADORES FINANCIEROS')
    lines.append('-' * 70)
    if result.indicators_result_json:
        try:
            from tendermod.evaluation.schemas import IndicatorComplianceResult
            ind = IndicatorComplianceResult.model_validate_json(result.indicators_result_json)
            cumple_str = 'CUMPLE' if ind.cumple else ('NO CUMPLE' if ind.cumple is False else 'INDETERMINADO')
            lines.append(f'Resultado: {cumple_str}')
            lines.append('')
            lines.append('Indicadores evaluados:')
            for nombre in ind.indicadores_evaluados:
                lines.append(f'  - {nombre}')
            lines.append('')
            lines.append('Argumentacion del LLM:')
            lines.append(ind.detalle)
        except Exception as exc:
            lines.append(f'Error al parsear: {exc}')
    else:
        lines.append('Sin datos de indicadores.')
    lines.append('')

    # Experiencia
    lines.append('-' * 70)
    lines.append('EXPERIENCIA RUP')
    lines.append('-' * 70)
    if result.experience_result_json:
        try:
            from tendermod.evaluation.schemas import ExperienceComplianceResult
            exp = ExperienceComplianceResult.model_validate_json(result.experience_result_json)
            cumple_str = 'CUMPLE' if exp.cumple else 'NO CUMPLE'
            lines.append(f'Resultado: {cumple_str}')
            lines.append(f'Codigos requeridos  : {", ".join(exp.codigos_requeridos)}')
            if exp.objeto_requerido:
                lines.append(f'Objeto requerido    : {exp.objeto_requerido}')
            if exp.objeto_exige_relevancia:
                lines.append(f'Exige relevancia    : {exp.objeto_exige_relevancia}')
            if exp.cantidad_contratos_requerida is not None:
                lines.append(f'Contratos (top-N)   : {exp.cantidad_contratos_requerida}')
            if exp.valor_requerido_cop:
                lines.append(f'Valor requerido     : ${int(round(exp.valor_requerido_cop)):,} COP'.replace(',', '.'))
            if exp.total_valor_cop:
                lines.append(f'Valor total acreditado: ${int(round(exp.total_valor_cop)):,} COP'.replace(',', '.'))
            lines.append(f'RUPs candidatos     : {len(exp.rups_candidatos_codigos)} contratos')
            lines.append(f'RUPs que cumplen    : {exp.rups_cumplen}')
            if exp.rups_excluidos_por_objeto:
                lines.append(f'RUPs excluidos por objeto: {exp.rups_excluidos_por_objeto}')
            lines.append('')
            lines.append('Detalle por RUP:')
            for rup in exp.rups_evaluados:
                cumple_rup = 'CUMPLE' if rup.cumple_total else 'NO CUMPLE'
                valor_str = f'${int(round(rup.valor_cop)):,} COP'.replace(',', '.') if rup.valor_cop else 'N/A'
                # FIX BUG-2: reemplazar \n en cliente
                cliente_str = (rup.cliente or 'N/A').replace('\n', ' ').replace('\r', '').strip()

                lines.append(f'  ┌─ RUP {rup.numero_rup} — {cliente_str} — {valor_str} — [{cumple_rup}]')
                lines.append(f'  │  Codigos UNSPSC  : {"CUMPLE" if rup.cumple_codigos else "NO CUMPLE"}')
                if rup.cumple_valor is not None:
                    lines.append(f'  │  Valor           : {"CUMPLE" if rup.cumple_valor else "NO CUMPLE"}')
                else:
                    lines.append('  │  Valor           : N/A')

                # FIX BUG-3: mostrar score aunque cumple_objeto sea None
                if rup.score_objeto is not None:
                    umbral_val = exp.similarity_threshold_usado
                    umbral_str = f'SUPERA umbral {umbral_val}' if rup.score_objeto >= umbral_val else f'BAJO umbral {umbral_val}'
                    if rup.cumple_objeto is False:
                        estado_obj = f'EXCLUIDO — Score {rup.score_objeto:.3f} ({umbral_str})'
                    elif rup.cumple_objeto is True:
                        estado_obj = f'CUMPLE — Score {rup.score_objeto:.3f} ({umbral_str})'
                    else:
                        estado_obj = f'Score {rup.score_objeto:.3f} ({umbral_str}) — filtro inactivo'
                    lines.append(f'  │  Objeto          : {estado_obj}')
                else:
                    obj_razon = 'sin datos en ChromaDB' if rup.cumple_objeto is False else 'N/A'
                    lines.append(f'  │  Objeto          : {obj_razon}')

                # FIX BUG-4: mostrar objeto_contrato
                if rup.objeto_contrato:
                    contrato_short = rup.objeto_contrato[:120].replace('\n', ' ')
                    if len(rup.objeto_contrato) > 120:
                        contrato_short += '...'
                    lines.append(f'  │  Contrato elegido: {contrato_short}')

                lines.append(f'  └─ Resultado      : {cumple_rup}')
                lines.append('')

            # Sub-requisitos MULTI_CONDICION (nivel global, no por RUP)
            if exp.sub_requisitos_resultado:
                lines.append('Sub-requisitos (MULTI_CONDICION):')
                for sr in exp.sub_requisitos_resultado:
                    estado = 'CUMPLE' if sr.cumple else 'NO CUMPLE'
                    score_str = f'Score={sr.score_objeto:.3f}' if sr.score_objeto is not None else 'Score=N/A'
                    rup_str = f'RUP={sr.rup_elegido}' if sr.rup_elegido is not None else 'RUP=Ninguno'
                    lines.append(f'  [{estado}] {sr.descripcion} ({score_str}, {rup_str})')
                lines.append('')
        except Exception as exc:
            lines.append(f'Error al parsear: {exc}')
    else:
        lines.append('Sin datos de experiencia.')

    lines.append('')
    lines.append('=' * 70)
    lines.append('Generado por tendermod')
    lines.append('=' * 70)

    content = '\n'.join(lines)
    filename = f'tendermod_resultados_{pk}.txt'
    response = HttpResponse(content, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_context(request, pk):
    """
    Descarga el contexto RAW del retriever usado para la evaluacion (indicadores + experiencia).
    """
    session = get_object_or_404(AnalysisSession, pk=pk)

    try:
        result = session.result
    except AnalysisResult.DoesNotExist:
        messages.error(request, 'No hay resultados para exportar.')
        return redirect('analysis:results', pk=pk)

    lines = []
    lines.append('=' * 70)
    lines.append('TENDERMOD — CONTEXTO DEL RETRIEVER (RAG)')
    lines.append('=' * 70)
    lines.append(f'Sesion ID   : {session.pk}')
    lines.append(f'PDF         : {session.pdf_filename}')
    lines.append(f'Fecha       : {session.created_at.strftime("%Y-%m-%d %H:%M")}')
    lines.append('')
    lines.append('-' * 70)
    lines.append('CONTEXTO — INDICADORES FINANCIEROS')
    lines.append('-' * 70)
    lines.append(result.indicators_context_text or '(sin contexto disponible — use el flujo PDF completo)')
    lines.append('')
    lines.append('-' * 70)
    lines.append('CONTEXTO — EXPERIENCIA')
    lines.append('-' * 70)
    lines.append(result.experience_context_text or '(sin contexto disponible — use el flujo PDF completo)')
    lines.append('')
    lines.append('=' * 70)

    content = '\n'.join(lines)
    filename = f'tendermod_contexto_rag_{pk}.txt'
    response = HttpResponse(content, content_type='text/plain; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ---------------------------------------------------------------------------
# Vista: Evaluacion Automatica
# ---------------------------------------------------------------------------

def analysis_auto(request, pk):
    """Renderiza la pagina de evaluacion automatica (observador del progreso backend)."""
    session = get_object_or_404(AnalysisSession, pk=pk)
    context = {
        'session': session,
        'auto_steps': AUTO_FLOW_STEPS,
        'results_url': reverse('analysis:results', kwargs={'pk': pk}),
        'progress_url': reverse('analysis:auto_progress', kwargs={'pk': pk}),
    }
    return render(request, 'analysis/auto.html', context)


@require_GET
def auto_progress(request, pk):
    """
    Devuelve el progreso del flujo automatico reconstruido desde la BD
    (timing_json + status). No depende del result backend de Celery, asi que
    recargar o volver a la pagina repinta el estado real en curso.
    """
    session = get_object_or_404(AnalysisSession, pk=pk)

    try:
        timing = json.loads(session.timing_json or '[]')
    except (json.JSONDecodeError, TypeError):
        timing = []

    # Mapear timing por nombre de paso (ultima entrada gana)
    timing_by_name = {e.get('paso'): e for e in timing if e.get('paso')}

    last_step_name = AUTO_FLOW_STEPS[-1][0]
    has_error = session.status == 'error' or any(e.get('estado') == 'error' for e in timing)
    finished = not has_error and (
        session.status == 'completed'
        or last_step_name in timing_by_name
        or (not session.auto_flow_active and bool(timing))
    )

    steps = []
    running_assigned = False
    total_s = 0.0
    done_weight = 0
    total_weight = sum(peso for _, _, peso in AUTO_FLOW_STEPS)

    for nombre, label, peso in AUTO_FLOW_STEPS:
        entry = timing_by_name.get(nombre)
        if entry is not None:
            estado = 'error' if entry.get('estado') == 'error' else 'ok'
            dur = entry.get('duracion_s')
            if dur:
                total_s += dur
            if estado == 'ok':
                done_weight += peso
            steps.append({'label': label, 'peso': peso, 'estado': estado, 'duracion_s': dur})
        else:
            # Primer paso sin entrada → es el que esta corriendo (salvo finished/error)
            if not running_assigned and not finished and not has_error:
                estado = 'running'
                running_assigned = True
            else:
                estado = 'pending'
            steps.append({'label': label, 'peso': peso, 'estado': estado, 'duracion_s': None})

    progreso_pct = round(done_weight / total_weight * 100) if total_weight else 0
    if finished:
        progreso_pct = 100

    return JsonResponse({
        'steps': steps,
        'progreso_pct': progreso_pct,
        'total_s': round(total_s, 1),
        'finished': finished,
        'error': has_error,
    })


@require_POST
def auto_evaluate_experience(request, pk):
    """Lanza evaluacion de experiencia usando los datos ya extraidos en la sesion."""
    session = get_object_or_404(AnalysisSession, pk=pk)
    exp_dict = json.loads(session.experience_requirements_json or '{}')
    threshold = SystemConfig.get_solo().threshold_objeto
    task = evaluate_experience_task.delay(session.id, exp_dict, threshold)
    return JsonResponse({'task_id': task.id})


@require_POST
def auto_evaluate_indicators(request, pk):
    """Lanza evaluacion de indicadores usando los datos ya extraidos en la sesion."""
    session = get_object_or_404(AnalysisSession, pk=pk)
    raw = json.loads(session.indicators_requirements_json or '{"answer": []}')
    indicators_list = raw.get('answer', [])
    task = evaluate_indicators_task.delay(session.id, indicators_list)
    return JsonResponse({'task_id': task.id})


@require_POST
def auto_evaluate_team_profiles(request, pk):
    """Lanza evaluacion de perfiles de equipo usando los datos ya extraidos en la sesion."""
    session = get_object_or_404(AnalysisSession, pk=pk)
    if not session.team_profiles_json:
        return JsonResponse({'error': 'No hay perfiles de equipo extraidos'}, status=400)
    task = evaluate_team_profiles_task.delay(session.id)
    return JsonResponse({'task_id': task.id})


@require_POST
def auto_generate_conclusion(request, pk):
    """Lanza generacion de conclusion ejecutiva usando los resultados ya calculados."""
    session = get_object_or_404(AnalysisSession, pk=pk)
    task = generate_conclusion_task.delay(session.id)
    return JsonResponse({'task_id': task.id})


# ---------------------------------------------------------------------------
# Vista: Evaluacion Rapida
# ---------------------------------------------------------------------------

def analysis_quick(request):
    """Redirige al nuevo chat Redneet (compatibilidad con links existentes)."""
    return redirect('analysis:quick_redneet')


def redneet_qa(request):
    """Chat conversacional unificado: contratos, equipo e indicadores de Redneet."""
    history = request.session.get('redneet_chat_history', [])
    return render(request, 'analysis/redneet_qa.html', {'chat_history': history})


@require_POST
def redneet_qa_query(request):
    """Endpoint síncrono: pregunta → ask_redneet → respuesta con memoria de conversación."""
    history = request.session.get('redneet_chat_history', [])
    try:
        body = json.loads(request.body)
        question = body.get('question', '').strip()
        if not question:
            return JsonResponse({'error': 'Pregunta vacía'}, status=400)
        from tendermod.evaluation.redneet_inference import ask_redneet
        answer = ask_redneet(question, chat_history=history)
        history.append({'role': 'user', 'content': question})
        history.append({'role': 'assistant', 'content': answer})
        request.session['redneet_chat_history'] = history[-10:]  # ventana: 5 turnos
        request.session.modified = True
        return JsonResponse({'answer': answer})
    except Exception as exc:
        logger.error('redneet_qa_query error: %s', exc)
        return JsonResponse({'error': str(exc)}, status=500)


@require_POST
def redneet_qa_clear(request):
    """Limpia el historial de conversación Redneet de la sesión Django."""
    request.session.pop('redneet_chat_history', None)
    request.session.modified = True
    return JsonResponse({'status': 'ok'})


@require_POST
def analysis_quick_evaluate(request):
    """
    POST (JSON): Lanza evaluaciones rapidas de experiencia o indicadores desde texto libre.
    Mantenido para compatibilidad con el flujo automatico del pliego.
    """
    try:
        body = json.loads(request.body)
        action = body.get('action')
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Body JSON invalido'}, status=400)

    if action == 'reset':
        request.session.pop('quick_session_id', None)
        return JsonResponse({'status': 'ok'})

    if action not in ('experience', 'indicators'):
        return JsonResponse({'error': f'Accion desconocida: {action}'}, status=400)

    plain_text = body.get('text', '').strip()
    if not plain_text:
        return JsonResponse({'error': 'El campo text no puede estar vacio'}, status=400)

    quick_session_id = request.session.get('quick_session_id')
    if quick_session_id:
        try:
            session = AnalysisSession.objects.get(pk=quick_session_id)
        except AnalysisSession.DoesNotExist:
            quick_session_id = None

    if not quick_session_id:
        session = AnalysisSession.objects.create(
            pdf_filename='[Evaluacion Rapida]',
            status='pdf_ready',
        )
        request.session['quick_session_id'] = session.pk

    if action == 'experience':
        try:
            umbral = float(body.get('umbral', SystemConfig.get_solo().threshold_objeto))
        except (TypeError, ValueError):
            umbral = float(SystemConfig.get_solo().threshold_objeto)
        task = quick_evaluate_experience_task.delay(session.pk, plain_text, umbral)
    else:
        task = quick_evaluate_indicators_task.delay(session.pk, plain_text)

    session.celery_task_id = task.id
    session.save(update_fields=['celery_task_id', 'updated_at'])
    return JsonResponse({'task_id': task.id, 'session_id': session.pk})


@require_GET
def download_ocr(request, pk):
    """Descarga el documento Word generado por OCR para una sesión."""
    from pathlib import Path
    session = get_object_or_404(AnalysisSession, pk=pk)
    if not session.ocr_document_path:
        raise Http404("Esta sesión no tiene documento OCR generado.")
    ocr_path = Path(session.ocr_document_path)
    if not ocr_path.exists():
        raise Http404("El archivo OCR no existe en el servidor.")
    return FileResponse(
        open(ocr_path, 'rb'),
        as_attachment=True,
        filename=ocr_path.name,
        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document',
    )


def download_pdf(request, pk):
    """Descarga el PDF de licitación ingresado en la sesión."""
    from pathlib import Path
    session = get_object_or_404(AnalysisSession, pk=pk)
    if not session.pdf_filename:
        raise Http404("Esta sesión no tiene PDF asociado.")
    pdf_path = Path(settings.TENDERMOD_DATA_DIR) / session.pdf_filename
    if not pdf_path.exists():
        raise Http404("El archivo PDF no existe en el servidor.")
    return FileResponse(
        open(pdf_path, 'rb'),
        as_attachment=True,
        filename=session.pdf_filename,
        content_type='application/pdf',
    )


@require_POST
def session_rename(request, pk):
    """Actualiza el display_name de una sesión (AJAX)."""
    session = get_object_or_404(AnalysisSession, pk=pk)
    try:
        body = json.loads(request.body)
        name = body.get('name', '').strip()
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'error': 'Body JSON inválido'}, status=400)
    session.display_name = name
    session.save(update_fields=['display_name', 'updated_at'])
    return JsonResponse({'ok': True, 'name': name or session.pdf_filename})


def team_qa(request):
    """Página de Evaluación Equipo: chat en lenguaje natural contra SQLite."""
    history = request.session.get('team_chat_history', [])
    return render(request, 'analysis/team_qa.html', {'chat_history': history})


@require_POST
def team_qa_query(request):
    """Endpoint síncrono: pregunta → pipeline team_inference → respuesta con memoria."""
    history = request.session.get('team_chat_history', [])
    try:
        body = json.loads(request.body)
        question = body.get('question', '').strip()
        if not question:
            return JsonResponse({'error': 'Pregunta vacía'}, status=400)
        from tendermod.evaluation.team_inference import ask_team
        answer = ask_team(question, chat_history=history)
        history.append({'role': 'user', 'content': question})
        history.append({'role': 'assistant', 'content': answer})
        request.session['team_chat_history'] = history[-10:]  # ventana: 5 turnos
        request.session.modified = True
        return JsonResponse({'answer': answer})
    except Exception as exc:
        logger.error('team_qa_query error: %s', exc)
        return JsonResponse({'error': str(exc)}, status=500)


@require_POST
def team_qa_clear(request):
    """Limpia el historial de conversación del equipo de la sesión Django."""
    request.session.pop('team_chat_history', None)
    request.session.modified = True
    return JsonResponse({'status': 'ok'})


@require_POST
def analysis_delete(request, pk):
    """Elimina una sesión y limpia todos sus datos: BD, PDF en disco y OCR Word."""
    from pathlib import Path
    session = get_object_or_404(AnalysisSession, pk=pk)

    # PDF original en disco (puede no existir si fue sobreescrito por otra sesión)
    if session.pdf_filename:
        (Path(settings.TENDERMOD_DATA_DIR) / session.pdf_filename).unlink(missing_ok=True)

    # Documento OCR Word si existe
    if session.ocr_document_path:
        Path(session.ocr_document_path).unlink(missing_ok=True)

    nombre = session.pdf_filename or f'sesión {pk}'
    session.delete()  # CASCADE elimina AnalysisResult automáticamente

    messages.success(request, f'Evaluación "{nombre}" eliminada correctamente.')
    return redirect('analysis:list')
